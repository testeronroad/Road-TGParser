from __future__ import annotations

import csv
from io import BytesIO, StringIO

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

EXPORT_HEADERS_RU = (
    "ID записи",
    "ID Telegram",
    "Название",
    "Описание",
    "Подписчики",
    "Ссылка",
    "Username",
    "Категория (код)",
    "Категория",
    "Ключ поиска",
    "Тип",
    "Дата парсинга",
)


def _row_values(row: dict) -> tuple:
    is_ch = row.get("is_broadcast") in (1, True)
    type_label = "Канал" if is_ch else "Супергруппа"
    return (
        row.get("id"),
        row.get("telegram_id"),
        row.get("title") or "",
        row.get("description") or "",
        row.get("members_count") if row.get("members_count") is not None else "",
        row.get("link") or "",
        row.get("username") or "",
        row.get("category_id") or "",
        row.get("category_name") or "",
        row.get("search_keyword") or "",
        type_label,
        row.get("parsed_at") or "",
    )


def build_csv_bytes(rows: list[dict]) -> bytes:
    buf = StringIO()
    writer = csv.writer(buf, delimiter=";", quoting=csv.QUOTE_MINIMAL)
    writer.writerow(EXPORT_HEADERS_RU)
    for row in rows:
        writer.writerow(_row_values(row))
    text = buf.getvalue()
    return ("\ufeff" + text).encode("utf-8")


def build_xlsx_bytes(rows: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Каналы"
    header_font = Font(bold=True)
    for col, h in enumerate(EXPORT_HEADERS_RU, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
    for r, row in enumerate(rows, start=2):
        for c, val in enumerate(_row_values(row), start=1):
            ws.cell(row=r, column=c, value=val)
    for col in range(1, len(EXPORT_HEADERS_RU) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 40
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()
